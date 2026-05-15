# Primeiro script desenvolvido na UNICOM - tirar dados de uma planilha e inserir em outras em PYTHON de forma automatica (com melhorias para o futuro)


import pandas as pd
# variavel mes
#################################################################
mes = pd.to_datetime("01/01/2025").strftime("%Y-%m") # formatação para sempre pegar na cèlula o mes correto
# janeiro:
# column_qnt = "G"
# column_valor = "H"
# column_nf = "I"
#################################################################
# # pegar planilha resumida
# # pegar nome das empresas coluna A e a linha q o nome está
# # abrir planilha individual
# # para cada nome de empresa na array de empresas
#     # pega a aba da planilha da empresa no arquivo individual
#     # procurar a linha do mês de janeiro na coluna B
#     # pegar valor na coluna D da mesma linha
# #  colcoar valor na resumida coluna G linha mesma Duarte

from openpyxl import load_workbook


# abrir planilha (utilizando workbook) mantendo formatação
wb_resume = load_workbook("sua planilha")
ws_resume = wb_resume["Plan1"]

# pegar planilha resumida
plan_resume = pd.read_excel("sua planilha", sheet_name="Plan1", header=None)


empresas_nome = []

# Percorrer coluna A
for indice, valor in plan_resume.iloc[:, 0].items():

    # Ignorar nulos
    if pd.isna(valor):
        continue

    texto = str(valor).strip()

    # Ignorar vazios
    if texto == "":
        continue

    empresas_nome.append({
        # pandas começa em 0
        # excel começa em 1
        "linha": indice + 1,
        "nome": texto
    })


# achar coluna do mês
coluna_mes = None

# Percorrer linha 1
for coluna, data in plan_resume.iloc[0].items():

    # Ignorar nulos
    if pd.isna(data):
        continue

    data = str(data).strip()

    # Ignorar vazios
    if data == "":
        continue

    # procurar mês
    if mes[:7] in data:
        # pandas começa em 0
        # openpyxl começa em 1
        coluna_mes = coluna + 1
        break


# abrir planilha individual
# para cada nome de empresa na array de empresas
for obj_empresa in empresas_nome:

    # print(f"Empresa: {obj_empresa['nome']}") para saber qual empresa 

    try:
        # pegar a aba da planilha da empresa no arquivo individual
        plan_indi = pd.read_excel(
            "sua planilha",
            sheet_name=obj_empresa["nome"]
        )
    except:
        print(f"Aba não encontrada: {obj_empresa['nome']}")
        continue


    # procurar a linha do mês de janeiro na coluna B
    linha_mes = None

    # Percorrer coluna B
    for indice2, valor2 in plan_indi.iloc[:, 1].items():

        # Ignorar nulos
        if pd.isna(valor2):
            continue

        texto2 = str(valor2).strip()

        # Ignorar vazios
        if texto2 == "":
            continue

        # procurar mês
        if mes[:7] in texto2:
            linha_mes = indice2
            break


    # se não encontrou mês
    if linha_mes is None:
        print("Mês não encontrado")
        continue


    # pegar valor qnt na coluna D (3)
    QNT = plan_indi.iloc[linha_mes, 3]

    # pegar valor na coluna E (4)
    VALOR = plan_indi.iloc[linha_mes, 4]

    # pegar nf na coluna F (5)
    NF = plan_indi.iloc[linha_mes, 5]


    # CNPJ (coluna A)
    cnpj = None
    for v in plan_indi.iloc[:, 0]:
        if pd.notna(v) and str(v).strip() != "":
            cnpj = str(v).strip()
            break


    # colocar valor na resumida
    # coluna G/H/I dependendo do mês
    linha_excel = obj_empresa["linha"]

    # QNT
    ws_resume.cell(
        row=linha_excel,
        column=coluna_mes,
        value=QNT
    )

    # VALOR
    ws_resume.cell(
        row=linha_excel,
        column=coluna_mes + 1,
        value=VALOR
    )

    # NF
    ws_resume.cell(
        row=linha_excel,
        column=coluna_mes + 2,
        value=NF
    )

    # CNPJ (vou colocar na coluna B da planilha resumo)
    ws_resume.cell(
        row=linha_excel,
        column=2,
        value=cnpj
    )

    print(f"{obj_empresa['nome']} processado com sucesso")


# salvar alterações sem perder formatação (tudo por conta da biblioteca openyxl)
wb_resume.save("Fatura Mensal VDO.xlsx")

print(plan_resume.head())
print("Dados inseridos com sucesso.")
