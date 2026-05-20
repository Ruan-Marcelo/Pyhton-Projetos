from copy import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# variavel mes
#################################################################

# panda para pegar todos os meses do ano e assim colocar na planilha dinamicamente
# Gerar meses de 2025 e 2026
meses = (
    pd.date_range(
        start="2025-01-01",
        end="2026-12-01",  # vai até dezembro de 2026
        freq="MS",  # MS = Month Start
    )
    .strftime("%Y-%m")
    .tolist()
)

#################################################################

# abrir planilha (utilizando workbook) mantendo formatação
wb_resume = load_workbook("")
ws_resume = wb_resume["Plan1"]

# pegar planilha resumida
plan_resume = pd.read_excel("", sheet_name="Plan1", header=None)


empresas_nome = []

# Percorrer coluna A
for indice, valor in plan_resume.iloc[:, 0].items():

    # Ignorar nulos
    if pd.isna(valor):
        continue

    texto = str(valor).strip()

    # Ignorar vazios
    if texto == "":
        continue

    empresas_nome.append(
        {
            # pandas começa em 0
            # excel começa em 1
            "linha": indice + 1,
            "nome": texto,
        }
    )


# abrir planilha individual
# para cada nome de empresa na array de empresas
for obj_empresa in empresas_nome:

    try:
        # pegar a aba da planilha da empresa no arquivo individual
        plan_indi = pd.read_excel(
            "", sheet_name=obj_empresa["nome"]
        )
    except Exception as e:
        print(f"Empresa não encontrada: {obj_empresa['nome']}")
        continue

    ### NOVO: Pegar o texto da célula A2 da planilha individual ###
    # No pandas com header=True (padrão), a linha 2 do Excel vira a linha 0.
    # Para garantir o comportamento independente de headers, usamos o read_excel padrão:
    # A célula A2 do Excel estará na linha de índice 0, coluna 0 se a linha 1 virou header.
    # Porém, olhando a sua imagem da planilha individual, a linha 1 possui "Irmãos Silva...",
    # a linha 2 possui "UNICOM DEPOSITO".
    # Se o pandas leu a linha 1 como cabeçalho, a linha 2 do Excel vira a linha 0 do DataFrame.
    # Para evitar problemas com o cabeçalho, acessamos o valor sabendo que 'UNICOM DEPOSITO' está em A2:

    try:
        # Lendo novamente sem header para garantir o local exato da célula A2 (Linha 2, Coluna A -> índice 1, 0)
        plan_indi_crua = pd.read_excel(
            "",
            sheet_name=obj_empresa["nome"],
            header=None,
        )
        texto_a2 = plan_indi_crua.iloc[1, 0]  # Linha 2 (índice 1), Coluna A (índice 0)
        texto_a2 = str(texto_a2).strip() if pd.notna(texto_a2) else ""
    except Exception as e:
        texto_a2 = ""
        print(f"Erro ao ler Nome da empresa {obj_empresa['nome']}: {e}")

    linha_excel = obj_empresa["linha"]

    # unicom ou uniktrom
    ###Colocar o texto de A2 na planilha resumida (Coluna B, linha da empresa) 
    # Coluna B no openpyxl é a coluna 2
    if texto_a2:
        celula_b = ws_resume.cell(row=linha_excel, column=2, value=texto_a2)
        # Copiar formatação da linha de cima 
        try:
            celula_b._style = copy(
                ws_resume.cell(row=linha_excel - 1, column=2)._style
            )
        except:
            pass

    ### NOVO: Pegar valor da célula C6 da planilha individual ###
    # Linha 6, Coluna C -> índice [5,2] no pandas
    try:
        valor_c6 = plan_indi_crua.iloc[5, 2]  # Linha 6, Coluna C
        valor_c6 = str(valor_c6).strip() if pd.notna(valor_c6) else ""
    except Exception as e:
        valor_c6 = ""
        print(f"Erro ao ler o calor da empresa {obj_empresa['nome']}: {e}")

    # Colocar esse valor na planilha resumida na coluna 5 (E) a partir da linha da empresa
    if valor_c6:
        celula_d = ws_resume.cell(row=linha_excel, column=5, value=valor_c6)  # Coluna E = 5
        # Copiar formatação da linha de cima 
        try:
            celula_d._style = copy(
                ws_resume.cell(row=linha_excel - 1, column=5)._style
            )
        except:
            pass

    # PERCORRER TODOS OS MESES
    for mes in meses:

        # achar coluna do mês
        coluna_mes = None

        # Percorrer linha 1
        for coluna, data in plan_resume.iloc[0].items():

            # Ignorar nulos
            if pd.isna(data):
                continue

            data = str(data).strip()

            # Ignorar vazios
            if data == "":
                continue

            # procurar mês
            if mes[:7] in data:
                # pandas começa em 0
                # openpyxl começa em 1
                coluna_mes = coluna + 1
                break

        # se não encontrou coluna do mês
        if coluna_mes is None:
            print(f"Coluna do mês não encontrada: {mes}")
            continue

        # procurar a linha do mês de janeiro na coluna B
        linha_mes = None

        # Percorrer coluna B
        for indice2, valor2 in plan_indi.iloc[:, 1].items():

            # Ignorar nulos
            if pd.isna(valor2):
                continue

            texto2 = str(valor2).strip()

            # Ignorar vazios
            if texto2 == "":
                continue

            # procurar mês
            if mes[:7] in texto2:
                linha_mes = indice2
                break

        # se não encontrou mês
        if linha_mes is None:
            print(f"Mês não encontrado: {mes}")
            continue

        # pegar valor qnt na coluna D (3)
        QNT = plan_indi.iloc[linha_mes, 3]

        # pegar valor na coluna E (4)
        VALOR = plan_indi.iloc[linha_mes, 4]

        # pegar nf na coluna F (5)
        NF = plan_indi.iloc[linha_mes, 5]

        # CNPJ (coluna A)
        cnpj = plan_indi.iloc[linha_mes, 0]

        # QNT
        celula_qnt = ws_resume.cell(row=linha_excel, column=coluna_mes, value=QNT)

        # copiar formatação
        celula_qnt._style = copy(
            ws_resume.cell(row=linha_excel - 1, column=coluna_mes)._style
        )

        # VALOR
        celula_valor = ws_resume.cell(
            row=linha_excel, column=coluna_mes + 1, value=VALOR
        )

        # largura fixa da coluna VALOR
        ws_resume.column_dimensions[get_column_letter(coluna_mes + 1)].width = 13

        # copiar formatação
        celula_valor._style = copy(
            ws_resume.cell(row=linha_excel - 1, column=coluna_mes + 1)._style
        )

        # NF
        celula_nf = ws_resume.cell(row=linha_excel, column=coluna_mes + 2, value=NF)

        # copiar formatação
        celula_nf._style = copy(
            ws_resume.cell(row=linha_excel - 1, column=coluna_mes + 2)._style
        )

        # CNPJ (coluna C da planilha resumo)
        celula_cnpj = ws_resume.cell(row=linha_excel, column=3, value=cnpj)
        # copiar formatação
        celula_cnpj._style = copy(
            ws_resume.cell(row=linha_excel - 1, column=3)._style
        )

        print(f"{obj_empresa['nome']} {mes} processado com sucesso")


# salvar alterações sem perder formatação
wb_resume.save("")

print("Dados inseridos com sucesso.")

