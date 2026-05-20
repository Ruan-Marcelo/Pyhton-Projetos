from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import copy
import difflib
import unicodedata

# ---------- Função para normalizar nomes ----------
def normalizar(texto):
    """Remove acentos, transforma em maiúsculas e remove espaços extras"""
    if texto:
        texto = texto.upper().strip()
        texto = ''.join(c for c in unicodedata.normalize('NFD', texto)
                        if unicodedata.category(c) != 'Mn')
        return texto
    return ""

# ---------- Função para limpar palavras indesejadas ----------
def limpar_palavras(texto, palavras_remover=None):
    """Remove palavras indesejadas de um texto"""
    if not texto:
        return ""
    texto = str(texto)
    if palavras_remover is None:
        palavras_remover = ["PIX", "BOLETO", "DEPOSITO", "DESCONTO EM FOLHA"]
    for palavra in palavras_remover:
        texto = texto.replace(palavra, "")
    return texto.strip()

# ---------- Abrir planilha principal ----------
arquivo = "../UNICOM COMERCIAL DE AUTO PEÇAS LTDA.xlsx"
wb = load_workbook(arquivo)

# Aba Faturamento
if "Faturamento" in wb.sheetnames:
    ws = wb["Faturamento"]
else:
    raise ValueError("A aba 'Faturamento' não foi encontrada no arquivo.")

# Dicionário para armazenar dados únicos por empresa
dados = defaultdict(lambda: {'placas': 0, 'codigo_sap': None})

# Ler dados a partir da linha 2
for row in ws.iter_rows(min_row=2):
    empresa = row[3].value  # coluna D
    codigo_sap = row[2].value  # coluna C
    placa = row[6].value  # coluna G

    if empresa:
        try:
            placa_valor = int(placa)
        except (TypeError, ValueError):
            placa_valor = 1

        dados[empresa]['placas'] += placa_valor
        if not dados[empresa]['codigo_sap'] and codigo_sap:
            dados[empresa]['codigo_sap'] = codigo_sap

# ---------- Criar/selecionar aba "Resumo" ----------
if "Resumo" in wb.sheetnames:
    ws_resumo = wb["Resumo"]
    # Limpar linhas antigas a partir da linha 3
    if ws_resumo.max_row >= 3:
        ws_resumo.delete_rows(3, ws_resumo.max_row - 2)
else:
    ws_resumo = wb.create_sheet("Resumo")

# Copiar estilos da linha 2
estilos = []
for col in ['A','B','C','D']:
    cell = ws_resumo[f'{col}2']
    estilos.append({
        'font': copy.copy(cell.font),
        'fill': copy.copy(cell.fill),
        'border': copy.copy(cell.border),
        'alignment': copy.copy(cell.alignment),
        'number_format': copy.copy(cell.number_format)
    })

# Data atual
data_hoje = datetime.now().date()

# Preencher dados a partir da linha 3
linha = 3
for empresa, info in dados.items():
    ws_resumo[f'A{linha}'] = empresa
    ws_resumo[f'B{linha}'] = info['codigo_sap'] if info['codigo_sap'] else ""
    ws_resumo[f'C{linha}'] = data_hoje
    ws_resumo[f'D{linha}'] = info['placas']

    # Aplicar estilos copiados da linha 2
    for i, col in enumerate(['A','B','C','D']):
        cell = ws_resumo[f'{col}{linha}']
        cell.font = estilos[i]['font']
        cell.fill = estilos[i]['fill']
        cell.border = estilos[i]['border']
        cell.alignment = estilos[i]['alignment']
        cell.number_format = estilos[i]['number_format']

    linha += 1

# ---------- Abrir planilha VDO ----------
arquivo_vdo = "../Fatura Mensal VDO 2026 - Copia.xlsx"
wb_vdo = load_workbook(arquivo_vdo)
if "GERAL FTS" in wb_vdo.sheetnames:
    ws_vdo = wb_vdo["GERAL FTS"]
else:
    raise ValueError("A aba 'GERAL FTS' não foi encontrada no arquivo VDO.")

estilo_e = estilos[3]  # usar estilo da coluna D para quantidade

# ---------- Preparar lista de nomes VDO limpos e normalizados ----------
nomes_vdo = [ws_vdo[f'B{r}'].value for r in range(3, ws_vdo.max_row + 1) if ws_vdo[f'B{r}'].value]
# Limpar palavras indesejadas e depois normalizar
nomes_vdo_norm = [normalizar(limpar_palavras(n)) for n in nomes_vdo]

# ---------- Preencher Quantidades VDO e Nome Encontrado ----------
for row in range(3, ws_resumo.max_row + 1):
    nome_empresa = ws_resumo[f'A{row}'].value
    nome_empresa_norm = normalizar(nome_empresa)
    
    quantidade_vdo = 0
    nome_encontrado = ""

    # Correspondência aproximada
    matches = difflib.get_close_matches(nome_empresa_norm, nomes_vdo_norm, n=1, cutoff=0.7)
    if matches:
        indice = nomes_vdo_norm.index(matches[0])
        nome_encontrado = nomes_vdo[indice]
        valor = ws_vdo[f'AY{indice + 3}'].value  # +3 porque começa na linha 3
        try:
            quantidade_vdo = int(valor)
        except (TypeError, ValueError):
            quantidade_vdo = 0

    # Preencher coluna E (quantidade)
    ws_resumo[f'E{row}'] = quantidade_vdo
    cell = ws_resumo[f'E{row}']
    cell.font = estilo_e['font']
    cell.fill = estilo_e['fill']
    cell.border = estilo_e['border']
    cell.alignment = estilo_e['alignment']
    cell.number_format = estilo_e['number_format']

    # Preencher coluna F (nome encontrado)
    ws_resumo[f'F{row}'] = nome_encontrado

# ---------- Salvar planilha ----------
arquivo_copia = ""
wb.save(arquivo_copia)
print("Planilha atualizada!")

# from openpyxl import load_workbook
# from datetime import datetime
# from collections import defaultdict
# import os
# import copy

# # Caminho do arquivo
# arquivo = "."

# # Abrir planilha e selecionar aba ""
# wb = load_workbook(arquivo)
# if "" in wb.sheetnames:
#     ws = wb[""]
# else:
#     raise ValueError("A aba '' não foi encontrada no arquivo.")

# # Dicionário para armazenar dados únicos por empresa
# dados = defaultdict(lambda: {'placas': 0, 'codigo_sap': None})

# # Ler dados a partir da linha 2
# for row in ws.iter_rows(min_row=2):
#     empresa = row[3].value  # coluna D
#     codigo_sap = row[2].value  # coluna C
#     placa = row[6].value  # coluna G

#     if empresa:
#         try:
#             placa_valor = int(placa)
#         except (TypeError, ValueError):
#             placa_valor = 1

#         dados[empresa]['placas'] += placa_valor
#         if not dados[empresa]['codigo_sap'] and codigo_sap:
#             dados[empresa]['codigo_sap'] = codigo_sap

# # Criar/selecionar aba "Resumo"
# if "Resumo" in wb.sheetnames:
#     ws_resumo = wb["Resumo"]
#     # Limpar linhas antigas a partir da linha 3
#     if ws_resumo.max_row >= 3:
#         ws_resumo.delete_rows(3, ws_resumo.max_row - 2)
# else:
#     ws_resumo = wb.create_sheet("Resumo")

# # Copiar estilos da linha 2
# estilos = []
# for col in ['A','B','C','D']:
#     cell = ws_resumo[f'{col}2']
#     estilos.append({
#         'font': copy.copy(cell.font),
#         'fill': copy.copy(cell.fill),
#         'border': copy.copy(cell.border),
#         'alignment': copy.copy(cell.alignment),
#         'number_format': copy.copy(cell.number_format)
#     })

# # Data atual
# data_hoje = datetime.now().date()

# # Preencher dados a partir da linha 3
# linha = 3
# for empresa, info in dados.items():
#     ws_resumo[f'A{linha}'] = empresa
#     ws_resumo[f'B{linha}'] = info['codigo_sap'] if info['codigo_sap'] else ""
#     ws_resumo[f'C{linha}'] = data_hoje
#     ws_resumo[f'D{linha}'] = info['placas']

#     # Aplicar estilos copiados da linha 2
#     for i, col in enumerate(['A','B','C','D']):
#         cell = ws_resumo[f'{col}{linha}']
#         cell.font = estilos[i]['font']
#         cell.fill = estilos[i]['fill']
#         cell.border = estilos[i]['border']
#         cell.alignment = estilos[i]['alignment']
#         cell.number_format = estilos[i]['number_format']

#     linha += 1

# # Salvar como cópia
# arquivo_copia =""
# wb.save(arquivo_copia)
# print(f"Planilha atualizada!")
